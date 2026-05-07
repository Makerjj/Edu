from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.xdf_report.excel_template import render_report
from src.xdf_report.models import Problem, ReportRequest, Student, StudentProgress


def _make_request(
    tmp_path: Path,
    problem_queries: list[str],
    after_class_problem_queries: list[str] | None = None,
) -> ReportRequest:
    return ReportRequest(
        team_name="易生活102 C1",
        training_name="二分查找",
        problem_queries=problem_queries,
        template_path=Path("tests/fixtures/template.xlsx"),
        output_dir=tmp_path,
        after_class_problem_queries=after_class_problem_queries or [],
    )


def test_render_report_writes_expected_cells_from_template(tmp_path: Path) -> None:
    output_path = tmp_path / "result.xlsx"
    request = _make_request(tmp_path, ["找苹果", "摘梨"])
    problems = [
        Problem(problem_id="P1001", title="找苹果"),
        Problem(problem_id="P1002", title="摘梨"),
    ]
    rows = [
        StudentProgress(
            student=Student(uid="u1", username="s1", nickname="学生1"),
            completion_by_problem={"P1001": "✅", "P1002": ""},
        ),
        StudentProgress(
            student=Student(uid="u2", username="s2", nickname="学生2"),
            completion_by_problem={"P1001": "", "P1002": "✅"},
        ),
    ]
    after_class_problems = [
        Problem(problem_id="P2001", title="验证密码"),
        Problem(problem_id="P2002", title="逢7过"),
    ]

    render_report(
        request=_make_request(
            tmp_path,
            ["找苹果", "摘梨"],
            ["验证密码", "逢7过"],
        ),
        problems=problems,
        after_class_problems=after_class_problems,
        rows=rows,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert workbook.sheetnames == ["结果"]

    sheet = workbook["结果"]
    assert sheet["A3"].value == "学生1"
    assert sheet["A4"].value == "学生2"
    assert sheet["A5"].value is None
    assert sheet["F2"].value == "找苹果"
    assert sheet["G2"].value == "摘梨"
    assert sheet["H2"].value is None
    assert sheet["L2"].value == "验证密码"
    assert sheet["M2"].value == "逢7过"
    assert sheet["N2"].value is None
    assert sheet["F3"].value == "✅"
    assert sheet["G3"].value is None
    assert sheet["F4"].value is None
    assert sheet["G4"].value == "✅"
    assert sheet["L3"].value is None
    assert sheet["M3"].value is None
    assert sheet["L4"].value is None
    assert sheet["M4"].value is None


def test_render_report_extends_students_beyond_top_sample_block(tmp_path: Path) -> None:
    output_path = tmp_path / "result.xlsx"
    problems = [Problem(problem_id="P1001", title="找苹果")]
    rows = [
        StudentProgress(
            student=Student(uid=f"u{index}", username=f"s{index}", nickname=f"学生{index}"),
            completion_by_problem={"P1001": "✅" if index % 2 else ""},
        )
        for index in range(1, 7)
    ]

    render_report(
        request=_make_request(tmp_path, ["找苹果"]),
        problems=problems,
        after_class_problems=[],
        rows=rows,
        output_path=output_path,
    )

    workbook = load_workbook(output_path)

    assert workbook.sheetnames == ["结果"]

    sheet = workbook["结果"]
    assert sheet.max_row == 8
    assert [sheet.cell(row, 1).value for row in range(3, 9)] == [
        "学生1",
        "学生2",
        "学生3",
        "学生4",
        "学生5",
        "学生6",
    ]
    assert sheet["A9"].value is None
    assert sheet["A13"].value is None


def test_render_report_writes_after_class_data_in_after_class_region(tmp_path: Path) -> None:
    output_path = tmp_path / "result.xlsx"
    problems = [Problem(problem_id="P1001", title="找苹果")]
    after_class_problems = [
        Problem(problem_id="P2001", title="验证密码"),
        Problem(problem_id="P2002", title="逢7过"),
    ]
    rows = [
        StudentProgress(
            student=Student(uid="u1", username="s1", nickname="学生1"),
            completion_by_problem={"P1001": "✅"},
            after_class_completion_by_problem={"P2001": "✅", "P2002": ""},
        ),
        StudentProgress(
            student=Student(uid="u2", username="s2", nickname="学生2"),
            completion_by_problem={"P1001": ""},
            after_class_completion_by_problem={"P2001": "", "P2002": "✅"},
        ),
    ]

    render_report(
        request=_make_request(
            tmp_path,
            ["找苹果"],
            ["验证密码", "逢7过"],
        ),
        problems=problems,
        after_class_problems=after_class_problems,
        rows=rows,
        output_path=output_path,
    )

    sheet = load_workbook(output_path)["结果"]

    assert sheet["F2"].value == "找苹果"
    assert sheet["F3"].value == "✅"
    assert sheet["F4"].value is None
    assert sheet["L2"].value == "验证密码"
    assert sheet["M2"].value == "逢7过"
    assert sheet["L3"].value == "✅"
    assert sheet["M3"].value is None
    assert sheet["L4"].value is None
    assert sheet["M4"].value == "✅"


def test_render_report_leaves_after_class_region_blank_when_not_provided(
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "result.xlsx"
    problems = [Problem(problem_id="P1001", title="找苹果")]
    rows = [
        StudentProgress(
            student=Student(uid="u1", username="s1", nickname="学生1"),
            completion_by_problem={"P1001": "✅"},
        )
    ]

    render_report(
        request=_make_request(tmp_path, ["找苹果"]),
        problems=problems,
        after_class_problems=[],
        rows=rows,
        output_path=output_path,
    )

    sheet = load_workbook(output_path)["结果"]

    assert sheet["L2"].value is None
    assert sheet["M2"].value is None
    assert sheet["L3"].value is None
    assert sheet["M3"].value is None


def test_render_report_fills_classroom_performance_with_three_stars(tmp_path: Path) -> None:
    output_path = tmp_path / "result.xlsx"
    problems = [Problem(problem_id="P1001", title="找苹果")]
    rows = [
        StudentProgress(
            student=Student(uid="u1", username="s1", nickname="学生1"),
            completion_by_problem={"P1001": "✅"},
        ),
        StudentProgress(
            student=Student(uid="u2", username="s2", nickname="学生2"),
            completion_by_problem={"P1001": ""},
        ),
    ]

    render_report(
        request=_make_request(tmp_path, ["找苹果"]),
        problems=problems,
        after_class_problems=[],
        rows=rows,
        output_path=output_path,
    )

    sheet = load_workbook(output_path)["结果"]

    assert sheet["C3"].value == "🌟🌟🌟"
    assert sheet["D3"].value == "🌟🌟🌟"
    assert sheet["E3"].value == "🌟🌟🌟"
    assert sheet["C4"].value == "🌟🌟🌟"
    assert sheet["D4"].value == "🌟🌟🌟"
    assert sheet["E4"].value == "🌟🌟🌟"


def test_render_report_fills_attendance_with_checkmarks(tmp_path: Path) -> None:
    output_path = tmp_path / "result.xlsx"
    problems = [Problem(problem_id="P1001", title="找苹果")]
    rows = [
        StudentProgress(
            student=Student(uid="u1", username="s1", nickname="学生1"),
            completion_by_problem={"P1001": "✅"},
        ),
        StudentProgress(
            student=Student(uid="u2", username="s2", nickname="学生2"),
            completion_by_problem={"P1001": ""},
        ),
    ]

    render_report(
        request=_make_request(tmp_path, ["找苹果"]),
        problems=problems,
        after_class_problems=[],
        rows=rows,
        output_path=output_path,
    )

    sheet = load_workbook(output_path)["结果"]

    assert sheet["B3"].value == "✅"
    assert sheet["B4"].value == "✅"


def test_render_report_removes_merged_cells_from_last_student_row(tmp_path: Path) -> None:
    template_path = tmp_path / "merged-tail-template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "模板"

    for row in range(1, 14):
        for col in range(1, 14):
            sheet.cell(row, col).value = f"r{row}c{col}"

    for row in range(3, 14):
        sheet.cell(row, 1).value = f"模板学生{row}"

    sheet.merge_cells("C13:E13")
    sheet.merge_cells("F13:I13")
    sheet.merge_cells("L13:M13")
    workbook.save(template_path)

    output_path = tmp_path / "result.xlsx"
    rows = [
        StudentProgress(
            student=Student(uid=f"u{index}", username=f"s{index}", nickname=f"学生{index}"),
            completion_by_problem={"P1001": "✅" if index == 11 else ""},
        )
        for index in range(1, 12)
    ]

    request = ReportRequest(
        team_name="易生活102 C1",
        training_name="二分查找",
        problem_queries=["找苹果"],
        template_path=template_path,
        output_dir=tmp_path,
    )

    render_report(
        request=request,
        problems=[Problem(problem_id="P1001", title="找苹果")],
        after_class_problems=[],
        rows=rows,
        output_path=output_path,
    )

    result_sheet = load_workbook(output_path)["结果"]

    merged_ranges = {str(cell_range) for cell_range in result_sheet.merged_cells.ranges}
    assert "C13:E13" not in merged_ranges
    assert "F13:I13" not in merged_ranges
    assert "L13:M13" not in merged_ranges
    assert result_sheet["A13"].value == "学生11"
    assert result_sheet["F13"].value == "✅"
    assert result_sheet["G13"].value is None
