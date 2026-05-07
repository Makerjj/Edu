from copy import copy
from pathlib import Path

from openpyxl import load_workbook

IN_CLASS_START_COL = 6
AFTER_CLASS_START_COL = 12
CLASSROOM_PERFORMANCE_START_COL = 3
CLASSROOM_PERFORMANCE_END_COL = 5
THREE_STAR_TEXT = "🌟🌟🌟"
ATTENDANCE_COL = 2
ATTENDANCE_CHECKMARK = "✅"


def _copy_row_style(sheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = sheet.cell(source_row, col)
        target = sheet.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height


def _clear_value_region(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            sheet.cell(row, col).value = None


def _remove_merged_ranges_below(sheet, max_row: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.max_row > max_row:
            sheet.merged_cells.ranges.remove(merged_range)


def _remove_merged_ranges_in_student_region(sheet, start_row: int, end_row: int) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row >= start_row and merged_range.max_row <= end_row:
            sheet.merged_cells.ranges.remove(merged_range)


def _sample_student_end_row(sheet, student_start_row: int) -> int:
    row = student_start_row
    while row <= sheet.max_row and sheet.cell(row, 1).value not in (None, ""):
        row += 1
    return row - 1


def _write_problem_region(
    sheet,
    header_row: int,
    start_col: int,
    problems,
    rows,
    completion_attr: str,
    student_start_row: int,
) -> None:
    for index, problem in enumerate(problems, start=start_col):
        header = sheet.cell(header_row, index)
        header.value = problem.title
        if index > start_col:
            left = sheet.cell(header_row, start_col)
            header._style = copy(left._style)

    for offset, row in enumerate(rows):
        target_row = student_start_row + offset
        completion_by_problem = getattr(row, completion_attr)
        for problem_index, problem in enumerate(problems, start=start_col):
            sheet.cell(target_row, problem_index).value = (
                completion_by_problem.get(problem.problem_id) or None
            )


def _write_classroom_performance_region(sheet, student_start_row: int, rows) -> None:
    for offset, _row in enumerate(rows):
        target_row = student_start_row + offset
        for col in range(CLASSROOM_PERFORMANCE_START_COL, CLASSROOM_PERFORMANCE_END_COL + 1):
            sheet.cell(target_row, col).value = THREE_STAR_TEXT


def _write_attendance_region(sheet, student_start_row: int, rows) -> None:
    for offset, _row in enumerate(rows):
        target_row = student_start_row + offset
        sheet.cell(target_row, ATTENDANCE_COL).value = ATTENDANCE_CHECKMARK


def render_report(request, problems, after_class_problems, rows, output_path: Path) -> None:
    workbook = load_workbook(request.template_path)
    template_sheet = workbook[workbook.sheetnames[-1]]
    sheet = workbook.copy_worksheet(template_sheet)
    sheet.title = "结果"

    student_start_row = 3
    max_col = max(
        sheet.max_column,
        IN_CLASS_START_COL + len(problems) - 1,
        AFTER_CLASS_START_COL + len(after_class_problems) - 1,
    )
    clear_end_col = max(sheet.max_column, max_col)
    sample_student_end_row = _sample_student_end_row(sheet, student_start_row)
    sample_student_count = sample_student_end_row - student_start_row + 1
    extra_student_count = max(0, len(rows) - sample_student_count)

    if extra_student_count:
        insert_at = sample_student_end_row + 1
        sheet.insert_rows(insert_at, extra_student_count)
        for target_row in range(insert_at, insert_at + extra_student_count):
            _copy_row_style(sheet, sample_student_end_row, target_row, max_col)

    final_student_end_row = max(sample_student_end_row, student_start_row + len(rows) - 1)

    if sheet.max_row > final_student_end_row:
        sheet.delete_rows(final_student_end_row + 1, sheet.max_row - final_student_end_row)
    _remove_merged_ranges_below(sheet, final_student_end_row)
    _remove_merged_ranges_in_student_region(sheet, student_start_row, final_student_end_row)

    _clear_value_region(sheet, 2, 2, IN_CLASS_START_COL, clear_end_col)
    _clear_value_region(sheet, student_start_row, final_student_end_row, 1, clear_end_col)

    for offset, row in enumerate(rows):
        target_row = student_start_row + offset
        if target_row > template_sheet.max_row:
            _copy_row_style(sheet, student_start_row, target_row, max_col)
        sheet.cell(target_row, 1).value = row.student.nickname

    _write_classroom_performance_region(
        sheet=sheet,
        student_start_row=student_start_row,
        rows=rows,
    )
    _write_attendance_region(
        sheet=sheet,
        student_start_row=student_start_row,
        rows=rows,
    )

    _write_problem_region(
        sheet,
        header_row=2,
        start_col=IN_CLASS_START_COL,
        problems=problems,
        rows=rows,
        completion_attr="completion_by_problem",
        student_start_row=student_start_row,
    )
    _write_problem_region(
        sheet,
        header_row=2,
        start_col=AFTER_CLASS_START_COL,
        problems=after_class_problems,
        rows=rows,
        completion_attr="after_class_completion_by_problem",
        student_start_row=student_start_row,
    )

    for name in list(workbook.sheetnames):
        if name != "结果":
            del workbook[name]

    workbook.active = 0
    workbook.save(output_path)
